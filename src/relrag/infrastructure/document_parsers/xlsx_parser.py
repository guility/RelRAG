"""Parser for .xlsx (Excel)."""

import io
from pathlib import Path

from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from relrag.domain.value_objects import PropertyType
from relrag.infrastructure.document_parsers.base import ParseResult
from relrag.infrastructure.document_parsers.metadata_keys import (
    normalize_value_for_storage,
)


def parse_xlsx(data: bytes, filename: str | None = None) -> ParseResult:
    """Extract text from all cells and basic metadata from .xlsx."""
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except (InvalidFileException, BadZipFile) as e:
        raise ValueError(f"Invalid or corrupted xlsx file: {e}") from e
    parts: list[str] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                parts.append(" ".join(cells))
    text = "\n".join(parts) if parts else " "
    properties: dict[str, tuple[str, PropertyType]] = {}
    cp = wb.properties
    if cp.creator:
        properties["author"] = (normalize_value_for_storage(cp.creator), PropertyType.STRING)
    if cp.created:
        properties["created_date"] = (cp.created.isoformat(), PropertyType.DATE)
    if cp.modified:
        properties["modified_date"] = (cp.modified.isoformat(), PropertyType.DATE)
    if cp.title:
        properties["title"] = (normalize_value_for_storage(cp.title), PropertyType.STRING)
    if filename:
        p = Path(filename)
        properties["source_file_name"] = (normalize_value_for_storage(p.name), PropertyType.STRING)
        properties["source_file_type"] = (normalize_value_for_storage("xlsx"), PropertyType.STRING)
    wb.close()
    return ParseResult(text=text, properties=properties)
